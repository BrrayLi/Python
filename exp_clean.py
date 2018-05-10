#encoding=utf-8
'''
step one 
Read from excel,删除多余的字符:\n,空格
step two
解析结构，用字符串组或者链表来完成，提取参数的方法？看是否存在已有的代码可用  模块sqlparse 
结构可类似
判断同一层级下的条件，计算当前范围内的左右括号数量
main_exp,length,para1,para2,....

相似度判断规则
1、直接使用list.index查找，完全相同才能确认
2、判断各元素，table para condition，次序应为para table condition或者table para condition 
3、考虑是否需要判断下层层级

select  from
from    where
union   
and     
step three 
进行比较，根据长度，结构相似度
'''

def Get_Condition(argue1):
    return 'sda'
def separate(argue1,argue2):
    temp_list=[]
    index_begin=0
    index_end=argue1.find(argue2)
    while index_end!=-1:
        #是否需要增加条件，第一个出现的（必须比）早，最后一个出现的）必须比（晚
        #argue1[index_begin:index_end].find('(')<=argur1[index_begin:index_end].find(')')   && \
        #argue1[index_begin:index_end].rfind('(')<=argur1[index_begin:index_end].rfind(')')
        if argue1[index_begin:index_end].count('(')==argue1[index_begin:index_end].count(')') and \
            argue1[index_begin:index_end].find('(')<=argue1[index_begin:index_end].find(')')   and \
            argue1[index_begin:index_end].rfind('(')<=argue1[index_begin:index_end].rfind(')'):
            temp_list.append(argue1[index_begin:index_end].strip())        
            index_begin=index_end+len(argue2)
        index_end=argue1.find(argue2,index_end+1)    
        #os.system("pause")
    temp_list.append(argue1[index_begin:].strip())
    return temp_list
def find_max_index(para1, para2):
    '''
    用于返回满足para1中最大值的位置，对应的para2中相应位置的值
    ''' 
    index=[]
    max_value=max(para1)
    start=0

    #调试
    # print max_value


    while True:
        try:
            index.append(para2[para1.index(max_value,start)])
            start=para1.index(max_value,start)+1
        except:
            break           
    return index    
def SQL_standard(SQL):
    #用于格式化SQL，最终返回    
    sql=SQL.replace('\n','').replace('_x000D_','') #清除回车及excel特有的换行符
    sql=' '.join(sql.split())
    sql=sql.lower() 
    return sql
def Funcition_remove(sql):
    #step 1     判断是否开始有函数,并去除最外层的函数
    while True:
        if sql.find('decode')==0:
            sql=sql[7:sql.rfind(')')]
        elif sql.find('nvl')==0:
            sql=sql[4:sql.rfind(')')]
        elif sql.find('(')==0:
            sql=sql[1:sql.rfind(')')]
        else:
            break
    return sql
def Get_Information(sql):    
    #第一个select 即是第一层select  from 
    index_begin=sql.find('select')
    index_end=sql.find('from')
    while sql[index_begin:index_end].count('(')!=sql[index_begin:index_end].count(')'):
        index_end=sql.find('from',index_end+1)
        if index_end==-1:
            break
    para1=sql[index_begin+7:index_end-1]
    para_list=separate(para1,',').sort(key=lambda x:len(x))

    #第一个from  where
    index_begin=index_end
    index_end=sql.find('where')
    while sql[index_begin:index_end].count('(')!=sql[index_begin:index_end].count(')'):
        index_end=sql.find('where',index_end+1)
        if index_end==-1:
            break
    if  index_end==-1:  
        table1=sql[index_begin+5:]
    else    :
        table1=sql[index_begin+5:index_end-1]
        condition=sql[index_end+6:]
    table_list=separate(table1,',').sort(key=lambda x:len(x))
    condition_list=separate(condition,'and').sort(key=lambda x:len(x))
    

    return  [para_list,table_list,condition_list]
def Compare_sql_1(inst,example_sum):    
    #无完全相同项
    #先对比表格,取相同表格数最大的记录
    table_result=[]
    target_index=[]
    for index   in  range(len(example_sum)):

        #暴力调试
        # print index

        if  len(example_sum[index])> 1 and len(example_sum[index][1])==len(inst[1]):
            target_index.append(1)
        else:
            target_index.append(0)

        
    #暴力调试
    # print target_index,len(example_sum)

    target_index=find_max_index(target_index,range(len(example_sum)))
    
    #暴力调试
    # print target_index

    for  index  in  target_index:
        example_table=example_sum[index][1]
        table_num_same=0
        for  table_name in inst[1]:
            try:
                example_table.index(table_name) #逐个表查找
                table_num_same+=1                
            except:
                pass
        table_result.append(table_num_same)
    target_index=find_max_index(table_result,target_index)

    #暴力调试
    # print 'STEP ONE:',target_index
    # os.system("pause")
    

    para_count_sum=[]
    #判断参数值是否一致,example_sum[0]
    for index in target_index:
        try:
            example_sum[index].index(inst[0])
            para_count_sum.append(len(inst[0]))
        except:                
            #对比参数列表,example_sum[0]
            example_para=example_sum[index][0]
            para_count=0
            for i in range(len(inst[0])):
                # if inst[0][i]==example_para[i]:
                #     para_count=para_count+1
                try:
                    example_para.index(inst[0][i])
                    para_count=para_count+1
                except:
                    pass
            para_count_sum.append(para_count)
    target_index=find_max_index(para_count_sum,target_index)#取符合条件的Index

    #暴力调试
    # print 'STEP TWO:',target_index


    #验证条件记录,example_sum[2]，实测验证条件较为困难，可能需要进一步处理
    #是否需要取=、in之类的操作符进行排查，比较此字段是否一致，而非整个condition一致
    condition_count_sum=[]
    for index in target_index:
        try:
            example_sum[index].index(inst[2])
            condition_count_sum.append(len(inst[2]))
        except:
            example_condition=example_sum[index][2]
            condition_count=0
            for i in range(len(inst[2])):

                # if  inst[2][i]==example_para[i]:
                #     condition_count=condition+1                   
                try:
                    example_condition.index(inst[2][i])  #严格一致
                    condition_count=condition_count+1
                except:
                    pass
            condition_count_sum.append(condition_count)
    
    #暴力调试
    # print target_index,condition_count_sum

    target_index=find_max_index(condition_count_sum,target_index)
    return  target_index


import random
import sqlparse
import os 
import openpyxl
import sys
import copy



#存放格式化后的SQL表达式
sql_sum=[]  

# para_sum=[]     
# table_sum=[]
# condition_sum=[]

#存放分割后的SQL，第一列为para,第二列为tables，第三列为condition
example_sum=[]  
#存放对应的转换后的表达式
example_result=[]

#读取作为样本的SQL
xls_file=openpyxl.load_workbook('exp_clean_1.xlsx')
sheet=xls_file[xls_file.sheetnames[0]]
for index in range(2,sheet.max_row+1):
    sheet_index='I'+str(index)
    sql=str(sheet[sheet_index].value)
    sql=SQL_standard(sql)
    #增加判断，若存在select则需要提取信息，若不存在select则直接取sql，
    if sql.find('selcet')!=-1:  
        sql=Funcition_remove(sql)
        inst=Get_Information(sql)
    else:
        inst=sql
    sql_sum.append(sql)
    example_sum.append(inst)

    #该sql对应的替换后的表示式，L列
    sheet_index='L'+str(index)
    sql=str(sheet[sheet_index].value)
    result_list=SQL_standard(sql)
    example_result.append(result_list)
xls_file.close()

#调试
# print   len(example_result)


#读取需要处理的SQL，即需要找到转换后的表达式的SQL；处理逻辑跟上述一致
list_sum=[]
list_sql_sum=[]
xls_file=openpyxl.load_workbook('exp.xlsx')
sheet=xls_file[xls_file.sheetnames[0]]
for index in range(1,sheet.max_row+1): #索引从1开始，非0
    sheet_index='A'+str(index)

    sql=str(sheet[sheet_index].value)
#    print sql
    sql=SQL_standard(sql)
    #增加判断，若存在select则需要提取信息，若不存在select则直接取sql，
    if sql.find('selcet')!=-1:  
        sql=Funcition_remove(sql)
        inst=Get_Information(sql)
    else:
        inst=sql
    list_sum.append(inst)
    list_sql_sum.append(sql)




'''
print para_list,table_list,condition_list
para_list.sort(key=lambda x:len(x),reverse=True)
table_list.sort(key=lambda x:len(x),reverse=True)
condition_list.sort(key=lambda x:len(x),reverse=True)
print para_list,table_list,condition_list
print type([para_list,table_list,condition_list])

'''
'''debug
print 'para1:'+para1+'\n'
print 'para_list:',len(para_list),'\n',para_list
print 'tables:',len(table1),'\n',table1
print 'tables:',len(table_list),'\n',table_list
print '\n'
print 'condition:'+condition
print 'condition_list:',len(condition_list),'\n'
print condition_list
'''
#print 'select '+para1+' from '+tab le1+' where '+condition


#字符串比较版本1
#基本思路，优先级table、para、condition,其中condition仅作为验证
#判断依据，取一致数量最大的作为基准

#供参考使用的例子,example_sum
#example_sum.sort(key=lambda x:len(x[1]),reverse=True) #根据表个数进行排序
# inst=copy.deepcopy(list_sql_sum[50])         #当前需要处理的实例

#用于暴力调试


# print sql_sum[1][1]
# print para_sum[1]
# print table_sum[1]
# print condition_sum[1]
# print [para_sum[1],table_sum[1],condition_sum[1]]
# inst[1]=['tb_cm_serv a', 'tb_ba_msobj b']
#inst[0]=['0']
# print inst


result=[]
for inst in list_sum:
    target_index=[]
    try:    #判断是否存在一致的语句
        # example_sum.index('123fds')
        index=example_sum.index(inst)
        target_index=[index]
    except:
        # print 'not THE SAME !!\n'
        index=-1    
    if index==-1 and len(inst)>1:
        target_index=Compare_sql_1(inst,example_sum)


        #暴力调试
        # print 'STEP THREE:',target_index

    else:
        pass
    if len(target_index)==0:
        result.append('找不到相似的SQL')
    else:
        # print 'compare_result is ',target_index
        try:
            result.append(example_result[target_index[random.randint(0,len(target_index)-1)]])
        except:
            print 'compare_result is ',target_index

'''
设置最终输出格式
原始SQL;参数列表;涉及表格列表;参考的新函数;入参;入参;入参;.......
以;为分隔符，存放在csv或者txt文件中
'''

#存放example数据，备用
file_exp_result=open('test.txt','w')
for i in range(len(sql_sum)):
    #file_exp_result.write(sql+';'+para_list+';'+table_list+';'+target_function+';'+para1+';'+'para2'+';'+....)
    file_exp_result.write(sql_sum[i]+'\n')
    # file_exp_result.write(';'+str(len(example_sum[i][0])))
    # for  string in example_sum[i][0]:
    #     file_exp_result.write(';'+string)
    # file_exp_result.write(';'+str(len(example_sum[i][1])))
    # for  string in example_sum[i][1]:
    #     file_exp_result.write(';'+string)
    # file_exp_result.write(';'+str(len(example_sum[i][2])))
    # for  string in example_sum[i][2]:
    #     file_exp_result.write(';'+string)
    # file_exp_result.write('\n')    
file_exp_result.close()

#存放转换后的表达式及原sql，即存放结果
file_exp_result=open('result.txt','w')
for i in range(len(list_sql_sum)):
    file_exp_result.write(list_sql_sum[i]+';'+result[i]+'\n')
file_exp_result.close()



