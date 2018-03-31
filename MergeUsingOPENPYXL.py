#encoding=utf-8

#xlxs module
import openpyxl

# windows modules
import tkFileDialog
import tkMessageBox

import os

#Read the path of files Using tkfiledialog
source_filename=tkFileDialog.askopenfilenames(title='file',filetypes=[('excel','*.xls *.xlsx')])
if len(source_filename)==0:
    print u'无文件被选择'
    exit() 
filename_target=os.path.splitext(source_filename[0])[0]+'-Merge.xlsx'
print os.path.split(filename_target),os.listdir(os.path.split(filename_target)[0])
if os.path.split(filename_target)[1] in os.listdir(os.path.split(filename_target)[0]):
    print u'文件已存在！'
#    exit()

#Read data and Write data
file_write=openpyxl.Workbook()
file_write.save(filename_target)
w1=openpyxl.load_workbook(filename_target)
Sheet_target=w1[w1.sheetnames[0]]
target_row=2
for temp in source_filename:
    data_temp=openpyxl.load_workbook(filename=temp)
    sheet=data_temp[data_temp.sheetnames[0]]
    rows_temp=sheet.max_row
    cols_temp=sheet.max_column
    print data_temp,rows_temp,cols_temp,sheet,Sheet_target
    for x in range(2,rows_temp+1):
        for y in range(97,+cols_temp+97): #chr(97)='a'
            y=chr(y)
            i='%s%d'%(y,x)
            j='%s%d'%(y,target_row)
            #print sheet[i].value
            Sheet_target[j].value=sheet[i].value
            #print Sheet_target[j].value
        target_row=target_row+1
else:
    for y in range(97,97+cols_temp):
        y=chr(y)
        i='%s%d'%(y,1)
        Sheet_target[i].value=sheet[i].value
    data_temp.close()
w1.save(filename_target)


