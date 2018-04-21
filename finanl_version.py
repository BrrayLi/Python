
#encoding=utf-8

import xlrd
#xlxs module
import openpyxl
# windows modules
import tkFileDialog
import tkMessageBox
import gc  
import os
import time 
import sys  

reload(sys)  

sys.setdefaultencoding('utf8')    

#if read from xls,using this function
def Xls_Read(file,source_xls):
    print '%f%s' %(time.clock(),'-openxls_start')
    xls_temp=xlrd.open_workbook(filename=source_xls)
    source_sheet=xls_temp.sheets()[0]
    print '%f%s' %(time.clock(),'-openxls_end')
    for index in range(1,source_sheet.nrows):
        file.write( ','.join(['%s' % str(string) for string in source_sheet.row_values(index)])+'\n')
    print time.clock()
    del xls_temp,source_sheet
    gc.collect()
    return

def Xlsx_Read(file,source_xlsx):
    print '%s%s' % (str(time.clock()),'-openxlsx_begin')
    xlsx_temp=openpyxl.load_workbook(filename=source_xlsx,read_only=True)  
    print '%s%s' % (str(time.clock()),'-openxlsx_end1')
    source_sheet=xlsx_temp[xlsx_temp.sheetnames[0]]
    print '%s%s' % (str(time.clock()),'-openxlsx_end2')
    for index in source_sheet.iter_rows(min_row=1,max_row=source_sheet.max_row,max_col=source_sheet.max_column):
        file.write(','.join(['%s' % str(value.value) for value in index])+'\n')  
    print '%s%s' % (str(time.clock()),'-xlsxwrite_end')
    xlsx_temp.close()   
    del xlsx_temp,source_sheet       
    gc.collect()
    print '%s%s' % (str(time.clock()),'-functionxlsxread_end')
    return  

source_filename=tkFileDialog.askopenfilenames(title='file',filetypes=[('excel','*.xls *.xlsx')])
if len(source_filename)==0:
    print u'无文件被选择'
    exit() 
filename_target=os.path.splitext(source_filename[0])[0]+'-Merge.csv'
if os.path.split(filename_target)[1] in os.listdir(os.path.split(filename_target)[0]):
    print u'文件已存在！'
    exit()

file=open(filename_target,"a")
for temp in source_filename:
    print temp
    if os.path.splitext(temp)[1]=='.xls':
        Xls_Read(file,temp)
    else:
        Xlsx_Read(file,temp)
print '%f%s' % (time.clock(),'+')
file.close()
print time.clock()